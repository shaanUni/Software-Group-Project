from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .audit import log_audit_event
from .forms import TeamCreateForm, UserCreateForm, DepartmentCreateForm, ProjectCreateForm
from .models import Department, Team


def is_superadmin(user):
    return user.is_authenticated and user.is_superuser


def is_admin_user(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def _display_user_name(user):
    if not user:
        return "Unassigned"

    full_name = user.get_full_name().strip()
    return full_name or user.username or "Unassigned"


def _build_report_rows():
    team_queryset = (
        Team.objects.select_related("team_leader", "department")
        .annotate(member_count=Count("users", distinct=True))
        .order_by("department__department_name", "team_name", "team_id")
    )

    team_rows = [
        {
            "team_id": team.team_id,
            "name": team.team_name or f"Team {team.team_id}",
            "department": team.department.department_name or f"Department {team.department.department_id}",
            "leader": _display_user_name(team.team_leader),
            "member_count": team.member_count,
        }
        for team in team_queryset
    ]

    department_queryset = (
        Department.objects.annotate(team_count=Count("teams", distinct=True))
        .order_by("department_name", "department_id")
    )

    department_rows = [
        {
            "department_id": department.department_id,
            "name": department.department_name or f"Department {department.department_id}",
            "team_count": department.team_count,
        }
        for department in department_queryset
    ]

    summary = {
        "total_teams": len(team_rows),
        "teams_without_managers": sum(1 for team in team_rows if team["leader"] == "Unassigned"),
        "total_departments": len(department_rows),
    }

    return team_rows, department_rows, summary


def _build_pdf_response(team_rows, department_rows, summary):
    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportsTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=22,
        textColor=colors.HexColor("#132238"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "ReportsSubtitle",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        textColor=colors.HexColor("#526071"),
        spaceAfter=14,
    )
    section_style = ParagraphStyle(
        "ReportsSection",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=14,
        textColor=colors.HexColor("#132238"),
        spaceBefore=12,
        spaceAfter=8,
    )

    story = [
        Paragraph("Team Reports", title_style),
        Paragraph("Administrative overview of team coverage, leadership gaps, and department distribution.", subtitle_style),
    ]

    overview_table = Table(
        [
            [
                Paragraph("<b>Total Teams</b>", styles["BodyText"]),
                Paragraph("<b>Teams Without Managers</b>", styles["BodyText"]),
                Paragraph("<b>Departments</b>", styles["BodyText"]),
            ],
            [
                Paragraph(str(summary["total_teams"]), styles["Title"]),
                Paragraph(str(summary["teams_without_managers"]), styles["Title"]),
                Paragraph(str(summary["total_departments"]), styles["Title"]),
            ],
        ],
        colWidths=[2.6 * inch, 2.8 * inch, 2.0 * inch],
    )
    overview_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fbff")),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#132238")),
                ("LINEBELOW", (0, 0), (-1, 0), 1, colors.HexColor("#d7e3f4")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#d7e3f4")),
                ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d7e3f4")),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.extend([overview_table, Spacer(1, 14)])

    story.append(Paragraph("Teams by Department", section_style))
    department_table_data = [["Department", "Team Count"]]
    department_table_data.extend([[row["name"], row["team_count"]] for row in department_rows])
    if len(department_table_data) == 1:
        department_table_data.append(["No departments found", "-"])

    department_table = Table(department_table_data, colWidths=[5.0 * inch, 1.5 * inch], repeatRows=1)
    department_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4b99")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("ALIGN", (1, 1), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.extend([department_table, Spacer(1, 14)])

    story.append(Paragraph("All Teams", section_style))
    team_table_data = [["Team", "Department", "Leader", "Members"]]
    team_table_data.extend(
        [[row["name"], row["department"], row["leader"], row["member_count"]] for row in team_rows]
    )
    if len(team_table_data) == 1:
        team_table_data.append(["No teams found", "-", "-", "-"])

    team_table = Table(team_table_data, colWidths=[2.9 * inch, 2.6 * inch, 2.2 * inch, 0.9 * inch], repeatRows=1)
    team_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#132238")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("ALIGN", (3, 1), (3, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(team_table)

    document.build(story)

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="team-reports.pdf"'
    response.write(buffer.getvalue())
    return response


def _style_excel_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="1F4B99")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D7E2"),
        right=Side(style="thin", color="D0D7E2"),
        top=Side(style="thin", color="D0D7E2"),
        bottom=Side(style="thin", color="D0D7E2"),
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        sheet.column_dimensions[column_letter].width = min(max_length + 4, 40)

    sheet.freeze_panes = "A2"


def _build_excel_response(team_rows, department_rows, summary):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Total Teams", summary["total_teams"]])
    summary_sheet.append(["Teams Without Managers", summary["teams_without_managers"]])
    summary_sheet.append(["Departments", summary["total_departments"]])

    departments_sheet = workbook.create_sheet("Departments")
    departments_sheet.append(["Department", "Team Count"])
    for row in department_rows:
        departments_sheet.append([row["name"], row["team_count"]])

    teams_sheet = workbook.create_sheet("Teams")
    teams_sheet.append(["Team", "Department", "Leader", "Members"])
    for row in team_rows:
        teams_sheet.append([row["name"], row["department"], row["leader"], row["member_count"]])

    for sheet in (summary_sheet, departments_sheet, teams_sheet):
        _style_excel_sheet(sheet)

    summary_sheet["A1"].fill = PatternFill("solid", fgColor="132238")
    departments_sheet["A1"].fill = PatternFill("solid", fgColor="1F4B99")
    teams_sheet["A1"].fill = PatternFill("solid", fgColor="132238")

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    data = buffer.getvalue()
    response = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="team-reports.xlsx"'
    response["Content-Length"] = str(len(data))
    return response


@login_required
def team_list(request):
    teams = Team.objects.select_related("team_leader", "department")
    query = request.GET.get('q', '').strip()

    if query:
        teams = teams.filter(
            Q(team_name__icontains=query) |
            Q(department__department_name__icontains=query) |
            Q(team_leader__username__icontains=query) |
            Q(team_leader__first_name__icontains=query) |
            Q(team_leader__last_name__icontains=query)
        )

    return render(request, "teams/list.html", {"teams": teams, "query": query})


@login_required
@user_passes_test(is_superadmin)
def superadmin_team_create(request):
    if request.method == "POST":
        form = TeamCreateForm(request.POST)
        if form.is_valid():
            team = form.save()

            log_audit_event(
                user=request.user,
                action="CREATE",
                obj=team,
                description=f"Created team '{team}'.",
            )
            messages.success(request, "Team created successfully.")
            return redirect("team-list")
    else:
        form = TeamCreateForm()

    return render(request, "teams/team_form.html", {"form": form})


@login_required
def team_detail(request, team_id):
    team = get_object_or_404(
        Team.objects.select_related("team_leader", "department").prefetch_related("users"),
        pk=team_id,
    )

    users = team.users.all()

    return render(
        request,
        "teams/detail.html",
        {
            "team": team,
            "users": users,
        },
    )


@login_required
@user_passes_test(is_superadmin)
def team_edit_view(request, pk):
    team = get_object_or_404(Team, pk=pk)

    if request.method == "POST":
        form = TeamCreateForm(request.POST, instance=team)
        old_name = str(team)
        if form.is_valid():
            updated_team = form.save()
            log_audit_event(
                user=request.user,
                action="UPDATE",
                obj=updated_team,
                description=f"Updated team '{old_name}' to '{updated_team}'.",
            )
            messages.success(request, "Team updated successfully.")
            return redirect("team-list")
    else:
        form = TeamCreateForm(instance=team)

    return render(
        request,
        "teams/team_form.html",
        {
            "form": form,
            "page_title": "Edit Team",
            "page_subtitle": "Update the team details below.",
            "submit_text": "Save Changes",
            "is_edit": True,
            "team": team,
        },
    )


@login_required
@user_passes_test(is_superadmin)
def user_create_view(request):
    if request.method == "POST":
        form = UserCreateForm(request.POST)
        if form.is_valid():
            new_user = form.save()
            log_audit_event(
                user=request.user,
                action="CREATE",
                obj=new_user,
                description=f"Created user '{new_user}'.",
            )
            messages.success(request, "User created successfully.")
            return redirect("user-create")
    else:
        form = UserCreateForm()

    return render(request, "teams/user_create.html", {"form": form})


@login_required
@user_passes_test(is_admin_user)
def reports_view(request):
    team_rows, department_rows, summary = _build_report_rows()
    download_format = request.GET.get("format")

    if download_format == "pdf":
        return _build_pdf_response(team_rows, department_rows, summary)

    if download_format == "xlsx":
        return _build_excel_response(team_rows, department_rows, summary)

    return render(
        request,
        "teams/reports.html",
        {
            "team_rows": team_rows,
            "department_rows": department_rows,
            "summary": summary,
        },
    )


def organisation_view(request):
    teams = Team.objects.select_related("downstream_dependency", "department").all().order_by("team_name")

    team_data = [
        {
            "id": team.team_id,
            "name": team.team_name or f"Team {team.team_id}",
            "department": team.department.department_name if team.department else None,
            "members": team.employee_count(),
            "downstream_dependency_id": (
                team.downstream_dependency.team_id if team.downstream_dependency else None
            ),
        }
        for team in teams
    ]

    return render(
        request,
        "teams/organisation.html",
        {
            "teams": teams,
            "team_data": team_data,
        },
    )


@login_required
@user_passes_test(is_superadmin)
def department_create_view(request):
    if request.method == "POST":
        form = DepartmentCreateForm(request.POST)

        if form.is_valid():
            department = form.save()

            log_audit_event(
                user=request.user,
                action="CREATE",
                obj=department,
                description=f"Created department '{department}'.",
            )

            messages.success(request, "Department created successfully.")
            return redirect("team-list")
    else:
        form = DepartmentCreateForm()

    return render(
        request,
        "teams/simple_form.html",
        {
            "form": form,
            "page_title": "Create Department",
            "page_subtitle": "Add a new department.",
            "submit_text": "Create Department",
            "cancel_url": "team-list",
        },
    )


@login_required
@user_passes_test(is_superadmin)
def project_create_view(request):
    if request.method == "POST":
        form = ProjectCreateForm(request.POST)

        if form.is_valid():
            project = form.save()

            log_audit_event(
                user=request.user,
                action="CREATE",
                obj=project,
                description=f"Created project '{project}'.",
            )

            messages.success(request, "Project created successfully.")
            return redirect("team-list")
    else:
        form = ProjectCreateForm()

    return render(
        request,
        "teams/simple_form.html",
        {
            "form": form,
            "page_title": "Create Project",
            "page_subtitle": "Add a new project and assign it to a team.",
            "submit_text": "Create Project",
            "cancel_url": "team-list",
        },
    )